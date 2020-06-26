#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @FileName  :do_excel.py
# @create by :2020/6/25 11:49
# @author    :liuyuqing
import random
import openpyxl
from openpyxl import load_workbook


class DoExcel:

    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.wb = load_workbook(self.file_name)
        self.sheet = self.wb[self.sheet_name]
        self.max_row = self.sheet.max_row

    def get_data(self, mode='all'):
        test_data = []
        for i in range(2, self.max_row+1):
            data = {}
            data['case_id'] = self.sheet.cell(i, 1).value
            data['module'] = self.sheet.cell(i, 2).value
            data['title'] = self.sheet.cell(i, 3).value
            data['url'] = self.sheet.cell(i, 4).value
            data['method'] = self.sheet.cell(i, 5).value
            if isinstance(self.sheet.cell(i, 6).value, str):
                if self.sheet.cell(i, 6).value.find('${age}') != -1:
                    data['data'] = self.sheet.cell(i, 6).value.replace('${age}', str(random.randint(1, 30)))
            else:
                data['data'] = self.sheet.cell(i, 6).value
            data['result'] = self.sheet.cell(i, 7).value
            data['expected'] = self.sheet.cell(i, 8).value
            test_data.append(data)
        if mode == 'all':
            final_data = test_data
        else:
            final_data = []
            for item in test_data:
                if item['case_id'] in mode:
                    final_data.append(item)
        return final_data

    def write_data(self, row, col, value):
        self.sheet.cell(row, col).value = value
        self.wb.save(self.file_name)


if __name__ == '__main__':
    excel = DoExcel('test_case.xlsx', 'Sheet1')
    # excel.write_data(3, 7, "写入数据")
    data = excel.get_data([3])[0]['data']
    print(data)
    print(type(data))

